from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pyexcel as p
import os


def convert_to_xlsx(filename: str):
    if filename.split('.')[-1] == 'xls':
        p.save_book_as(file_name=filename,
                       dest_file_name=filename + 'x')
        return filename + 'x'


class ExcelBook:
    def __init__(self, filename: str):
        extension = filename.split('.')[-1].lower()

        if extension not in ['xlsx', 'xls']:
            raise AttributeError('File extension must be .xlsx or .xls')
        elif extension == 'xls':
            self.filename = convert_to_xlsx(filename)
        else:
            self.filename = filename

        if os.path.exists(self.filename):
            self.book = load_workbook(self.filename, data_only=True)
            self.number_of_sheets = len(self.book.worksheets)
        else:
            self.book = Workbook()
            self.number_of_sheets = 0

    def set_sheet_title(self, index, title):
        self.book.worksheets[index].title = title
        self.book.save(self.filename)
    
    def create_new_sheet(self, title=None):
        self.book.create_sheet(title=title)
        self.book.save(self.filename)


    def get_cell_value(self, cell: str, sheet_index=0, round_value=True):

        try:
            sheet = self.book.worksheets[sheet_index]
        except IndexError:
            return None

        # print(f"cell: {cell}")
        # value = sheet[cell].value
        value = sheet[cell].value if sheet[cell].value is not None else 0

        if type(value) is float and round_value:
            value = round(value, 2)

        return value

    def set_cell_value(
                self, cell, value, sheet_index=0, center=True, 
                fill_color: str|None = None, bold=False, rotate=False
        ):
        
        sheet = self.book.worksheets[sheet_index]

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        sheet[cell].value = value
        sheet[cell].border = thin_border

        if center:
            sheet[cell].alignment = Alignment(
                horizontal='center', vertical='center'
            )
        
        if fill_color is not None:
            pattern_fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type='solid'
            )
            sheet = self.book.worksheets[sheet_index]
            sheet[cell].fill = pattern_fill
        
        sheet[cell].font = Font(bold=bold)

        if rotate:
            sheet[cell].alignment = Alignment(text_rotation=90)

        self.book.save(self.filename)

    def get_max_row(self, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        return sheet.max_row

    def fill_cell(self, cell: str, fill_color: str, sheet_index=0):
        pattern_fill = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type='solid'
        )
        sheet = self.book.worksheets[sheet_index]
        sheet[cell].fill = pattern_fill
        self.book.save(self.filename)

        # ws[f'B{row}'].fill

    def set_column_width(self, column: str, width: int, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        sheet.column_dimensions[column].width = width
        self.book.save(self.filename)


def make_report_excel(report_dict, new_file_name):
    regions = [
        'Total',
        'Алмазарский р. (Ташкент)',
        'Бектемирский р. (Ташкент)',
        'Мирабадский р. (Ташкент)',
        'М.Улугбекский р. (Ташкент)',
        'Сергелийский р. (Ташкент)',
        'Чиланзарский р. (Ташкент)',
        'Шайхантахурский р. (Ташкент)',
        'Юнусабадский р. (Ташкент)',
        'Яккасарайский р. (Ташкент)',
        'Яшнабадский р. (Ташкент)',
        'Учтепинский р. (Ташкент)',
        'Андижанская обл.',
        'Бухарская обл.',
        'Джизакская обл.',
        'Кашкадарьинская обл.',
        'Навоинская обл.',
        'Наманганская обл.',
        'Респ. Каракалпакстан',
        'Самаркандская обл.',
        'Сурхандарьинская обл.',
        'Сырдарьинская обл.',
        'Ташкентская обл.',
        'Ферганская обл.',
        'Хорезмская обл.',
        'Неизвестный р.'
    ]

    wb = Workbook()

    fill_1 = PatternFill(
        start_color='F2FFAF',
        end_color='F2FFAF',
        fill_type='solid'
    )

    fill_1_1 = PatternFill(
        start_color='F8FFD7',
        end_color='F8FFD7',
        fill_type='solid'
    )

    fill_2 = PatternFill(
        start_color='BBFFC6',
        end_color='BBFFC6',
        fill_type='solid'
    )

    fill_2_1 = PatternFill(
        start_color='D5FFDC',
        end_color='D5FFDC',
        fill_type='solid'
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    region_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                      'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X',
                      'Y', 'Z', 'AA']

    for distributor in report_dict:

        ws = wb.create_sheet(distributor)
        wb.active = ws

        ws.merge_cells('A1:B1')
        ws.column_dimensions['B'].width = 40
        ws['A1'] = 'Наименование'
        ws['A1'].border = thin_border
        ws['B1'].border = thin_border
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(
            horizontal='center', vertical='center'
        )

        for i in range(0, len(regions)):
            ws[f'{region_columns[i]}1'] = regions[i]
            ws[f'{region_columns[i]}1'].border = thin_border
            ws[f'{region_columns[i]}1'].font = Font(bold=True)
            ws[f'{region_columns[i]}1'].alignment = \
                Alignment(text_rotation=90)

        group_row = 2
        group_index = 0
        for group in report_dict[distributor]:
            names_count = len(report_dict[distributor][group])
            ws.merge_cells(f'A{group_row}:A{group_row + names_count - 1}')
            ws[f'A{group_row}'] = group
            ws[f'A{group_row}'].font = Font(bold=True)
            ws[f'A{group_row}'].alignment = Alignment(
                text_rotation=90,
                horizontal='center',
                vertical='center'
            )

            ws[f'A{group_row}'].fill = fill_1 \
                if group_index % 2 == 0 \
                else fill_2

            group_row += names_count
            group_index += 1

        row = 2
        group_index = 0

        for group in report_dict[distributor]:
            for name in report_dict[distributor][group]:
                ws[f'B{row}'] = name

                ws[f'B{row}'].fill = fill_1_1 \
                    if group_index % 2 == 0 \
                    else fill_2_1

                ws[f'B{row}'].border = thin_border

                ws[f'A{row}'].border = thin_border

                for i in range(len(regions)):
                    sold_count = \
                        report_dict[distributor][group][name][regions[i]]

                    ws[f'{region_columns[i]}{row}'] = int(sold_count)

                    if float(sold_count) > int(sold_count):
                        ws[f'{region_columns[i]}{row}'].number_format = \
                            '# ##0.###'

                    ws[f'{region_columns[i]}{row}'].border = thin_border

                row += 1
            group_index += 1

    wb.remove(wb.worksheets[0])
    wb.save(new_file_name)



# print(str(ExcelBook("Альфа_вассерман_март.xls").get_cell_value("B18", sheet_index=0)))
# print(int(abs(ExcelBook("Соренто Апрель 2024.XLSX").get_cell_value("H3"))))




new_book = ExcelBook("Hello.xlsx")
new_book.set_cell_value("D5", value="SOS")





